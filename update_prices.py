import openpyxl
import asyncio
from pathlib import Path
import sys
sys.path.insert(0, '/home/ubuntu/icd10-search-engine')

async def update_prices():
    try:
        # Load Excel file
        wb = openpyxl.load_workbook('/home/ubuntu/upload/APPROVALFinalv15(3).xlsx')
        ws = wb.active
        
        # Import after path setup
        from server.db import getDb
        from drizzle.schema import drugLens
        from drizzle_orm import eq
        
        db = await getDb()
        updated = 0
        not_found = 0
        
        # Iterate through rows (skip header)
        for row_num in range(2, ws.max_row + 1):
            scientific_name = ws.cell(row_num, 1).value
            trade_name = ws.cell(row_num, 2).value
            price = ws.cell(row_num, 3).value
            
            if not trade_name or not price:
                continue
            
            # Try to find drug by trade name first
            result = await db.select().from_(drugLens).where(eq(drugLens.tradeName, trade_name)).limit(1)
            
            if not result and scientific_name:
                result = await db.select().from_(drugLens).where(eq(drugLens.scientificName, scientific_name)).limit(1)
            
            if result:
                # Update price
                await db.update(drugLens).set({'price': f'{price} SAR'}).where(eq(drugLens.id, result[0].id))
                updated += 1
                print(f"✓ Updated: {trade_name} - {price} SAR")
            else:
                not_found += 1
                print(f"✗ Not found: {trade_name}")
        
        print(f"\nSummary: {updated} updated, {not_found} not found")
    except Exception as e:
        print(f"Error: {str(e)}")

if __name__ == '__main__':
    asyncio.run(update_prices())
